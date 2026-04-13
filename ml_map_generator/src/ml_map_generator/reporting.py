from __future__ import annotations

import csv
import json
import re
import statistics
import xml.etree.ElementTree as ET
from collections import defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from .runner import RunResult

NUMBER_RE = re.compile(r"-?\d+")
TEXT_FILE_SUFFIXES = {'.log', '.txt', '.out', '.err', '.json', '.jsonl', '.jlog'}


@dataclass
class RunMetrics:
    run_id: str
    variant_label: str
    map_name: str
    size_group: str
    base_map: str
    family: str
    map_path: str
    repeat_index: int
    total_civilians: int
    total_search_buildings: int
    saved_civilians: int
    first_rescue_tick: Optional[int]
    last_rescue_tick: Optional[int]
    avg_ticks_between_rescues: Optional[float]
    rescued_percent: Optional[float]
    explored_buildings_percent: Optional[float]
    avg_decision_time_ms: Optional[float]
    avg_path_calc_time_ms: Optional[float]
    first_move_delay_ticks: Optional[float]
    overall_ticks: Optional[int]
    wall_runtime_sec: float
    repeatability_cv_percent: Optional[float]
    finished: bool
    notes: str


@dataclass
class SummaryRow:
    group_type: str
    group_name: str
    runs: int
    total_civilians_avg: Optional[float]
    total_search_buildings_avg: Optional[float]
    saved_civilians_avg: Optional[float]
    first_rescue_tick_avg: Optional[float]
    last_rescue_tick_avg: Optional[float]
    avg_ticks_between_rescues_avg: Optional[float]
    rescued_percent_avg: Optional[float]
    explored_buildings_percent_avg: Optional[float]
    avg_decision_time_ms_avg: Optional[float]
    avg_path_calc_time_ms_avg: Optional[float]
    first_move_delay_ticks_avg: Optional[float]
    overall_ticks_avg: Optional[float]
    wall_runtime_sec_avg: Optional[float]
    repeatability_cv_percent: Optional[float]
    saved_civilians_std: Optional[float]
    notes: str


@dataclass(frozen=True)
class RuntimeEvent:
    tick: int
    module_type: str
    event_type: str
    payload: dict[str, Any]
    wall_elapsed: Optional[float] = None
    source_file: str = ''


@dataclass(frozen=True)
class ReportBundle:
    session_dir: Path
    report_xlsx_path: Path
    report_csv_path: Path
    summary_csv_path: Path
    report_json_path: Path
    report_md_path: Path
    summary_md_path: Path


def build_variant_label(project_root: Path) -> str:
    runtime_dir = project_root / 'BaseRescueAgent' / 'config' / 'runtime'
    values: list[str] = []
    for name in ('detector', 'search', 'path'):
        path = runtime_dir / f'{name}.json'
        if not path.exists():
            continue
        try:
            data = json.loads(path.read_text(encoding='utf-8'))
        except Exception:
            continue
        section = data.get(name, {}) if isinstance(data, dict) else {}
        mode = str(section.get('mode', 'heuristic'))
        use_ml = bool(section.get('use_ml', False))
        values.append(f'{name}={mode}{"+ml" if use_ml else ""}')
    build_tag = None
    system_path = runtime_dir / 'system.json'
    if system_path.exists():
        try:
            data = json.loads(system_path.read_text(encoding='utf-8'))
            build_tag = ((data.get('system') or {}).get('build_tag') or (data.get('logs') or {}).get('build_tag'))
        except Exception:
            build_tag = None
    if build_tag:
        values.append(f'build={build_tag}')
    return ';'.join(values) if values else 'BaseRescueAgent'


def create_report_bundle(
    project_root: Path,
    run_results: list['RunResult'],
    variant_label: str,
    session_label: str,
) -> ReportBundle:
    table_root = project_root / 'BaseRescueAgent' / 'runtime' / 'table'
    table_root.mkdir(parents=True, exist_ok=True)
    session_dir = table_root / f'table_{session_label}'
    session_dir.mkdir(parents=True, exist_ok=True)

    run_metrics = [build_run_metrics(item, variant_label) for item in run_results]
    summary_rows = build_summary_rows(run_metrics)

    report_csv_path = session_dir / 'benchmark_runs.csv'
    summary_csv_path = session_dir / 'benchmark_summary.csv'
    report_md_path = session_dir / 'benchmark_runs.md'
    summary_md_path = session_dir / 'benchmark_summary.md'
    report_json_path = session_dir / 'benchmark_report.json'
    report_xlsx_path = session_dir / 'benchmark_report.xlsx'

    write_csv(report_csv_path, [asdict(item) for item in run_metrics])
    write_csv(summary_csv_path, [asdict(item) for item in summary_rows])
    write_markdown(report_md_path, [asdict(item) for item in run_metrics])
    write_markdown(summary_md_path, [asdict(item) for item in summary_rows])
    write_workbook(report_xlsx_path, run_metrics, summary_rows, session_dir)
    payload = {
        'created_at': datetime.now().isoformat(timespec='seconds'),
        'session_dir': str(session_dir),
        'variant_label': variant_label,
        'runs': [asdict(item) for item in run_metrics],
        'summary': [asdict(item) for item in summary_rows],
    }
    report_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')

    latest_path = table_root / 'latest_benchmark_report.xlsx'
    try:
        latest_path.write_bytes(report_xlsx_path.read_bytes())
    except Exception:
        pass

    return ReportBundle(
        session_dir=session_dir,
        report_xlsx_path=report_xlsx_path,
        report_csv_path=report_csv_path,
        summary_csv_path=summary_csv_path,
        report_json_path=report_json_path,
        report_md_path=report_md_path,
        summary_md_path=summary_md_path,
    )


def build_run_metrics(run_result: 'RunResult', variant_label: str) -> RunMetrics:
    map_dir = run_result.generated_dir / 'map'
    total_civilians, refuge_ids, building_ids = parse_map_metadata(map_dir)
    raw_events = load_runtime_events(run_result.raw_log_paths)
    debug_events = load_runtime_events(run_result.debug_log_paths)
    states = collapse_runtime_states(raw_events)
    rescue_ticks, saved_count, rescue_source = detect_rescues_from_runtime_states(states, refuge_ids)
    visited_buildings = detect_visited_buildings(states, building_ids)
    decision_ms = compute_average_decision_time_ms(raw_events, debug_events)
    path_ms = compute_average_path_calc_time_ms(raw_events)
    first_move_delay = compute_first_move_delay(states, raw_events)
    overall_ticks = max(
        [run_result.max_tick_seen]
        + [event.tick for event in raw_events]
        + [event.tick for event in debug_events]
        + [0]
    )
    rescued_percent = round((saved_count / total_civilians) * 100.0, 2) if total_civilians > 0 else None
    explored_percent = round((len(visited_buildings) / len(building_ids)) * 100.0, 2) if building_ids else None
    notes = [
        f'run_id={run_result.agent_run_id}',
        f'rescue_source={rescue_source}',
        f'copied_raw_logs={len(run_result.raw_log_paths)}',
        f'copied_debug_logs={len(run_result.debug_log_paths)}',
        f'copied_text_logs={len(run_result.text_log_paths)}',
    ]
    if run_result.failure_reason:
        notes.append(f'failure_reason={run_result.failure_reason}')
    if run_result.expected_timesteps is not None:
        notes.append(f'expected_timesteps={run_result.expected_timesteps}')
    if run_result.finished:
        notes.append('finished=true')

    return RunMetrics(
        run_id=run_result.map_name + '__r01',
        variant_label=variant_label,
        map_name=run_result.map_name,
        size_group=run_result.size_group,
        base_map=run_result.base_map,
        family=run_result.family,
        map_path=str(map_dir),
        repeat_index=1,
        total_civilians=total_civilians,
        total_search_buildings=len(building_ids),
        saved_civilians=saved_count,
        first_rescue_tick=rescue_ticks[0] if rescue_ticks else None,
        last_rescue_tick=rescue_ticks[-1] if rescue_ticks else None,
        avg_ticks_between_rescues=average_intervals(rescue_ticks),
        rescued_percent=rescued_percent,
        explored_buildings_percent=explored_percent,
        avg_decision_time_ms=round_or_none(decision_ms),
        avg_path_calc_time_ms=round_or_none(path_ms),
        first_move_delay_ticks=round_or_none(first_move_delay),
        overall_ticks=overall_ticks or None,
        wall_runtime_sec=round(run_result.wall_runtime_sec, 3),
        repeatability_cv_percent=None,
        finished=run_result.finished,
        notes='; '.join(notes),
    )


def load_runtime_events(paths: list[Path]) -> list[RuntimeEvent]:
    events: list[RuntimeEvent] = []
    for path in sorted(paths):
        if not path.exists() or path.suffix.lower() != '.jsonl':
            continue
        try:
            with path.open('r', encoding='utf-8', errors='ignore') as fh:
                for raw_line in fh:
                    line = raw_line.strip()
                    if not line:
                        continue
                    try:
                        record = json.loads(line)
                    except Exception:
                        continue
                    tick = to_int(record.get('tick'))
                    event_type = str(record.get('event_type') or '')
                    module_type = str(record.get('module_type') or '')
                    payload = record.get('payload') if isinstance(record.get('payload'), dict) else {}
                    if tick is None or not event_type:
                        continue
                    events.append(
                        RuntimeEvent(
                            tick=tick,
                            module_type=module_type,
                            event_type=event_type,
                            payload=payload,
                            wall_elapsed=to_float(record.get('wall_elapsed')),
                            source_file=path.name,
                        )
                    )
        except Exception:
            continue
    events.sort(key=lambda item: (item.tick, item.wall_elapsed if item.wall_elapsed is not None else -1.0, item.module_type, item.source_file))
    return events


def collapse_runtime_states(events: list[RuntimeEvent]) -> list[dict[str, Any]]:
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for event in events:
        if event.event_type != 'agent_runtime_state':
            continue
        payload = dict(event.payload)
        payload['_tick'] = event.tick
        payload['_module_type'] = event.module_type
        payload['_source_file'] = event.source_file
        grouped[event.tick].append(payload)
    collapsed: list[dict[str, Any]] = []
    for tick in sorted(grouped):
        candidates = grouped[tick]
        best = sorted(
            candidates,
            key=lambda item: (
                len(item.get('visited_buildings') or []),
                len(item.get('entered_buildings') or []),
                1 if item.get('_module_type') == 'search' else 0,
                1 if item.get('_module_type') == 'detector' else 0,
            ),
        )[-1]
        collapsed.append(best)
    return collapsed


def detect_rescues_from_runtime_states(states: list[dict[str, Any]], refuge_ids: set[int]) -> tuple[list[int], int, str]:
    if not states:
        return [], 0, 'none'
    rescue_ticks: list[int] = []
    rescued_ids: set[int] = set()
    previous: Optional[dict[str, Any]] = None
    for current in states:
        if previous is None:
            previous = current
            continue
        was_carrying = bool(previous.get('carrying'))
        now_carrying = bool(current.get('carrying'))
        if was_carrying and not now_carrying:
            civilian_id = to_int(previous.get('detector_target'))
            previous_position = to_int(previous.get('last_position'))
            current_position = to_int(current.get('last_position'))
            refuge_target = to_int(previous.get('refuge_target')) or to_int(current.get('refuge_target'))
            release_area = current_position or previous_position or refuge_target
            is_refuge_release = False
            if refuge_ids:
                is_refuge_release = release_area in refuge_ids or refuge_target in refuge_ids or (release_area is not None and refuge_target is not None and release_area == refuge_target)
            else:
                is_refuge_release = release_area is not None and refuge_target is not None and release_area == refuge_target
            if civilian_id is not None and is_refuge_release and civilian_id not in rescued_ids:
                rescued_ids.add(civilian_id)
                rescue_ticks.append(current.get('_tick') or previous.get('_tick') or 0)
        previous = current
    if not rescue_ticks:
        return [], 0, 'none'
    return rescue_ticks, len(rescued_ids), 'runtime_state_carrying_transition'


def detect_visited_buildings(states: list[dict[str, Any]], building_ids: set[int]) -> set[int]:
    visited: set[int] = set()
    for state in states:
        for value in state.get('visited_buildings') or []:
            number = to_int(value)
            if number is not None and (not building_ids or number in building_ids):
                visited.add(number)
        for value in state.get('entered_buildings') or []:
            number = to_int(value)
            if number is not None and (not building_ids or number in building_ids):
                visited.add(number)
    return visited


def compute_average_decision_time_ms(raw_events: list[RuntimeEvent], debug_events: list[RuntimeEvent]) -> Optional[float]:
    values: list[float] = []
    for event in raw_events:
        if event.event_type != 'decision_snapshot':
            continue
        state = event.payload.get('state') if isinstance(event.payload.get('state'), dict) else {}
        compute_ms = to_float(state.get('compute_ms'))
        if compute_ms is not None:
            values.append(compute_ms)
    for event in debug_events:
        if event.event_type != 'search_cycle_summary':
            continue
        compute_ms = to_float(event.payload.get('compute_ms'))
        if compute_ms is not None:
            values.append(compute_ms)
    if not values:
        return None
    return round(sum(values) / len(values), 3)


def compute_average_path_calc_time_ms(raw_events: list[RuntimeEvent]) -> Optional[float]:
    values: list[float] = []
    for event in raw_events:
        if event.event_type != 'path_snapshot':
            continue
        result = event.payload.get('result') if isinstance(event.payload.get('result'), dict) else {}
        compute_ms = to_float(result.get('compute_ms'))
        if compute_ms is not None:
            values.append(compute_ms)
    if not values:
        return None
    return round(sum(values) / len(values), 3)


def compute_first_move_delay(states: list[dict[str, Any]], raw_events: list[RuntimeEvent]) -> Optional[float]:
    first_goal_tick: Optional[int] = None
    for event in raw_events:
        if event.event_type == 'decision_snapshot':
            selected_id = event.payload.get('selected_id')
            if selected_id not in (None, '', 'None'):
                first_goal_tick = event.tick
                break
        if event.event_type == 'path_snapshot':
            request = event.payload.get('request') if isinstance(event.payload.get('request'), dict) else {}
            if request.get('to') not in (None, '', 'None'):
                first_goal_tick = event.tick
                break
    if first_goal_tick is None:
        return None
    first_move_tick: Optional[int] = None
    for state in states:
        move_tick = to_int(state.get('first_real_move_tick'))
        if move_tick is not None and move_tick >= 0:
            first_move_tick = move_tick
            break
    if first_move_tick is None:
        return None
    return float(max(first_move_tick - first_goal_tick, 0))


def parse_map_metadata(map_dir: Path) -> tuple[int, set[int], set[int]]:
    scenario_path = map_dir / 'scenario.xml'
    if not scenario_path.exists():
        raise FileNotFoundError(f'scenario.xml not found: {scenario_path}')
    root = ET.parse(scenario_path).getroot()
    refuge_ids: set[int] = set()
    civilian_ids: set[int] = set()
    synthetic_civilian_id = 1

    for elem in root.iter():
        tag = strip_ns(elem.tag).lower()
        blob = (tag + ' ' + ' '.join(f'{k}={v}' for k, v in elem.attrib.items()) + ' ' + (elem.text or '')).lower()
        entity_type = (elem.attrib.get('type') or elem.attrib.get('entity_type') or '').lower()
        is_refuge = 'refuge' in tag or entity_type == 'refuge' or 'refuge' in blob
        is_civilian = 'civilian' in tag or entity_type == 'civilian' or 'civilian' in blob

        if is_refuge:
            entity_id = first_int_attr(elem.attrib, ['id', 'entity_id', 'entityId', 'value', 'location'])
            location = first_int_attr(elem.attrib, ['location', 'position', 'area', 'id'])
            if location is not None:
                refuge_ids.add(location)
            elif entity_id is not None:
                refuge_ids.add(entity_id)

        if is_civilian:
            entity_id = first_int_attr(elem.attrib, ['id', 'entity_id', 'entityId', 'value'])
            if entity_id is None:
                entity_id = synthetic_civilian_id
                synthetic_civilian_id += 1
            civilian_ids.add(entity_id)

    building_ids, refuge_from_gml = parse_buildings_and_refuges_from_map_gml(map_dir)
    refuge_ids |= refuge_from_gml
    building_ids -= refuge_ids
    return len(civilian_ids), refuge_ids, building_ids


def parse_buildings_and_refuges_from_map_gml(map_dir: Path) -> tuple[set[int], set[int]]:
    candidates = [map_dir / 'map.gml'] + list(map_dir.glob('*.gml'))
    target = next((path for path in candidates if path.exists()), None)
    if target is None:
        return set(), set()
    building_ids: set[int] = set()
    refuge_ids: set[int] = set()
    try:
        root = ET.parse(target).getroot()
        for elem in root.iter():
            blob = (strip_ns(elem.tag) + ' ' + ' '.join(f'{k}={v}' for k, v in elem.attrib.items()) + ' ' + (elem.text or '')).lower()
            entity_id = first_int_attr(elem.attrib, ['id', 'entity_id', 'entityId', 'value', 'gml:id', 'fid'])
            if entity_id is None:
                entity_id = extract_first_int(blob)
            if entity_id is None:
                continue
            if 'refuge' in blob:
                refuge_ids.add(entity_id)
                continue
            if 'building' in blob or 'ambulancecentre' in blob or 'firestation' in blob or 'policeoffice' in blob:
                building_ids.add(entity_id)
    except Exception:
        pass
    if refuge_ids or building_ids:
        return building_ids, refuge_ids
    try:
        text = target.read_text(encoding='utf-8', errors='ignore')
    except Exception:
        return set(), set()
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        lower = line.lower()
        if not any(token in lower for token in ('refuge', 'building', 'ambulancecentre', 'firestation', 'policeoffice')):
            continue
        window = ' '.join(lines[max(0, idx - 2): idx + 3]).lower()
        entity_id = extract_first_int(window)
        if entity_id is None:
            continue
        if 'refuge' in window:
            refuge_ids.add(entity_id)
        elif any(token in window for token in ('building', 'ambulancecentre', 'firestation', 'policeoffice')):
            building_ids.add(entity_id)
    return building_ids, refuge_ids


def build_summary_rows(run_metrics: list[RunMetrics]) -> list[SummaryRow]:
    groups_by_map: dict[str, list[RunMetrics]] = defaultdict(list)
    groups_by_size: dict[str, list[RunMetrics]] = defaultdict(list)
    for item in run_metrics:
        groups_by_map[item.map_name].append(item)
        groups_by_size[item.size_group or 'unknown'].append(item)
    rows: list[SummaryRow] = []
    apply_repeatability_to_runs(run_metrics, groups_by_map)
    for map_name in sorted(groups_by_map):
        rows.append(build_summary_row('map', map_name, groups_by_map[map_name]))
    for size_group in ['small', 'medium', 'large', 'unknown']:
        if size_group in groups_by_size:
            rows.append(build_summary_row('size', size_group, groups_by_size[size_group]))
    if run_metrics:
        rows.append(build_summary_row('overall', 'OVERALL', run_metrics))
    return rows


def apply_repeatability_to_runs(run_metrics: list[RunMetrics], groups: dict[str, list[RunMetrics]]) -> None:
    by_map = {map_name: coefficient_of_variation([item.rescued_percent for item in items]) for map_name, items in groups.items()}
    for item in run_metrics:
        item.repeatability_cv_percent = by_map.get(item.map_name)


def build_summary_row(group_type: str, group_name: str, items: list[RunMetrics]) -> SummaryRow:
    note = 'summary_by_map' if group_type == 'map' else 'summary_by_size' if group_type == 'size' else 'overall_average'
    return SummaryRow(
        group_type=group_type,
        group_name=group_name,
        runs=len(items),
        total_civilians_avg=average_values([item.total_civilians for item in items]),
        total_search_buildings_avg=average_values([item.total_search_buildings for item in items]),
        saved_civilians_avg=average_values([item.saved_civilians for item in items]),
        first_rescue_tick_avg=average_values([item.first_rescue_tick for item in items]),
        last_rescue_tick_avg=average_values([item.last_rescue_tick for item in items]),
        avg_ticks_between_rescues_avg=average_values([item.avg_ticks_between_rescues for item in items]),
        rescued_percent_avg=average_values([item.rescued_percent for item in items]),
        explored_buildings_percent_avg=average_values([item.explored_buildings_percent for item in items]),
        avg_decision_time_ms_avg=average_values([item.avg_decision_time_ms for item in items]),
        avg_path_calc_time_ms_avg=average_values([item.avg_path_calc_time_ms for item in items]),
        first_move_delay_ticks_avg=average_values([item.first_move_delay_ticks for item in items]),
        overall_ticks_avg=average_values([item.overall_ticks for item in items]),
        wall_runtime_sec_avg=average_values([item.wall_runtime_sec for item in items]),
        repeatability_cv_percent=coefficient_of_variation([item.rescued_percent for item in items]),
        saved_civilians_std=std_values([item.saved_civilians for item in items]),
        notes=note,
    )


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open('w', encoding='utf-8', newline='') as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: empty_if_none(value) for key, value in row.items()})


def write_markdown(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    widths = {header: len(header) for header in headers}
    prepared: list[list[str]] = []
    for row in rows:
        current = [str(empty_if_none(row.get(header, ''))) for header in headers]
        prepared.append(current)
        for header, value in zip(headers, current):
            widths[header] = max(widths[header], len(value))
    with path.open('w', encoding='utf-8') as fh:
        header_line = '| ' + ' | '.join(header.ljust(widths[header]) for header in headers) + ' |\n'
        sep_line = '| ' + ' | '.join('-' * widths[header] for header in headers) + ' |\n'
        fh.write(header_line)
        fh.write(sep_line)
        for row in prepared:
            fh.write('| ' + ' | '.join(value.ljust(widths[header]) for value, header in zip(row, headers)) + ' |\n')


def write_workbook(path: Path, run_metrics: list[RunMetrics], summary_rows: list[SummaryRow], session_dir: Path) -> None:
    wb = Workbook()
    ws_runs = wb.active
    ws_runs.title = 'Runs'
    write_sheet(ws_runs, [asdict(item) for item in run_metrics])
    ws_summary = wb.create_sheet('Summary')
    write_sheet(ws_summary, [asdict(item) for item in summary_rows])
    ws_about = wb.create_sheet('About')
    ws_about['A1'] = 'variant_label'
    ws_about['B1'] = run_metrics[0].variant_label if run_metrics else ''
    ws_about['A2'] = 'session_dir'
    ws_about['B2'] = str(session_dir)
    ws_about['A3'] = 'notes'
    ws_about['B3'] = 'avg_decision_time_ms считается по compute_ms из decision_snapshot/search_cycle_summary; avg_path_calc_time_ms — по compute_ms из path_snapshot; size_group вынесен в отдельную колонку и summary_by_size.'
    style_sheet(ws_about)
    wb.save(path)


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws['A1'] = 'Нет данных'
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([empty_if_none(row.get(header)) for header in headers])
    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E1F2')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        values = ['' if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 40)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill('solid', fgColor='F7FBFF')


def strip_ns(tag: str) -> str:
    return tag.split('}')[-1] if '}' in tag else tag


def first_int_attr(attrs: dict[str, str], keys: Iterable[str]) -> Optional[int]:
    lowered = {str(key).lower(): value for key, value in attrs.items()}
    for key in keys:
        value = lowered.get(key.lower())
        number = to_int(value)
        if number is not None:
            return number
    return None


def extract_first_int(text: str) -> Optional[int]:
    match = NUMBER_RE.search(text)
    return int(match.group(0)) if match else None


def to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.lower() in {'none', 'null', ''}:
            return None
        match = NUMBER_RE.search(stripped)
        if match:
            return int(match.group(0))
    return None


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip().replace(',', '.')
        if stripped.lower() in {'none', 'null', ''}:
            return None
        try:
            return float(stripped)
        except Exception:
            match = re.search(r'-?\d+(?:\.\d+)?', stripped)
            if match:
                return float(match.group(0))
    return None


def average_values(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if not numeric:
        return None
    return round(sum(numeric) / len(numeric), 3)


def std_values(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if len(numeric) < 2:
        return None
    return round(statistics.pstdev(numeric), 3)


def coefficient_of_variation(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if len(numeric) < 2:
        return None
    mean = sum(numeric) / len(numeric)
    if abs(mean) < 1e-9:
        return None
    return round((statistics.pstdev(numeric) / mean) * 100.0, 3)


def average_intervals(values: list[int]) -> Optional[float]:
    if len(values) < 2:
        return None
    diffs = [values[index] - values[index - 1] for index in range(1, len(values))]
    return round(sum(diffs) / len(diffs), 3) if diffs else None


def round_or_none(value: Optional[float]) -> Optional[float]:
    return None if value is None else round(value, 3)


def empty_if_none(value: Any) -> Any:
    return '' if value is None else value
