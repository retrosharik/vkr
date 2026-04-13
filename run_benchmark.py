from __future__ import annotations

import argparse
import csv
import json
import math
import os
import re
import shutil
import signal
import statistics
import subprocess
import sys
import time
import traceback
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TICK_RE = re.compile(r"(?:^|\b)(?:tick|sim_elapsed|time|timestep|cycle|step)=(\d+)")
PREFIX_RE = re.compile(
    r"^\[module=(?P<module>[^\s\]]+) agent=(?P<agent>\d+) tick=(?P<tick>\d+) sim_elapsed=(?P<sim>\d+) "
    r"wall_elapsed=(?P<wall>[0-9.]+)s area=(?P<area>[^\s]+) x=(?P<x>[^\s]+) y=(?P<y>[^\]]+)\] (?P<msg>.*)$"
)
NUMBER_RE = re.compile(r"-?\d+")
EVENT_WORDS = (
    "saved",
    "rescued",
    "rescue_complete",
    "transport_complete",
    "victim_delivered",
    "delivered_to_refuge",
    "unload",
    "unloaded",
)
TEXT_FILE_SUFFIXES = {".log", ".txt", ".out", ".err", ".json", ".jsonl", ".jlog"}


@dataclass
class MapRun:
    name: str
    map_path: str
    config_path: str
    repeats: int = 1


@dataclass
class ModuleEvent:
    module: str
    agent: str
    tick: int
    sim_elapsed: int
    wall_elapsed_sec: float
    area: Optional[int]
    x: Optional[int]
    y: Optional[int]
    message: str
    source_file: str
    event_type: str
    target_id: Optional[int] = None
    state: Optional[str] = None


@dataclass
class RunMetrics:
    run_id: str
    variant_label: str
    map_name: str
    map_path: str
    repeat_index: int
    total_civilians: int
    total_search_buildings: int
    saved_civilians: int
    first_rescue_tick: Optional[int]
    last_rescue_tick: Optional[int]
    avg_ticks_between_rescues: Optional[float]
    rescued_percent: Optional[float]
    explored_buildings_percent: Optional[float]
    avg_decision_time_ms: Optional[float]
    avg_path_calc_time_ms: Optional[float]
    first_move_delay_ticks: Optional[float]
    overall_ticks: Optional[int]
    wall_runtime_sec: float
    repeatability_cv_percent: Optional[float]
    notes: str


@dataclass
class SummaryRow:
    group_name: str
    runs: int
    total_civilians_avg: Optional[float]
    total_search_buildings_avg: Optional[float]
    saved_civilians_avg: Optional[float]
    first_rescue_tick_avg: Optional[float]
    last_rescue_tick_avg: Optional[float]
    avg_ticks_between_rescues_avg: Optional[float]
    rescued_percent_avg: Optional[float]
    explored_buildings_percent_avg: Optional[float]
    avg_decision_time_ms_avg: Optional[float]
    avg_path_calc_time_ms_avg: Optional[float]
    first_move_delay_ticks_avg: Optional[float]
    overall_ticks_avg: Optional[float]
    wall_runtime_sec_avg: Optional[float]
    repeatability_cv_percent: Optional[float]
    saved_civilians_std: Optional[float]
    notes: str


class BenchmarkError(Exception):
    pass


class IncrementalLineReader:
    def __init__(self, path: Path) -> None:
        self.path = path
        self.position = 0
        self._buffer = ""

    def read_new_lines(self) -> list[str]:
        if not self.path.exists():
            return []
        with self.path.open("r", encoding="utf-8", errors="ignore") as fh:
            fh.seek(self.position)
            data = fh.read()
            self.position = fh.tell()
        if not data:
            return []
        data = self._buffer + data
        if data.endswith("\n"):
            self._buffer = ""
            lines = data.splitlines()
        else:
            lines = data.splitlines()
            self._buffer = lines.pop() if lines else data
        return lines


class BenchmarkRunner:
    def __init__(self, config: dict[str, Any]) -> None:
        self.config = config
        self.project_root = self._detect_project_root()
        self.server_scripts_dir = self._resolve_path(config.get("server_scripts_dir", "rcrs-server/scripts"))
        self.agent_dir = self._resolve_path(config.get("agent_dir", "newAgent"))
        self.agent_result_dir = self.agent_dir / "result"
        self.output_root = self._resolve_path(config.get("output_dir", "test"))
        self.output_root.mkdir(parents=True, exist_ok=True)
        self.label = str(config.get("run_label", "default"))
        self.variant_label = str(config.get("variant_label", self.label))
        self.startup_delay_sec = float(config.get("startup_delay_sec", 5))
        self.poll_interval_sec = float(config.get("poll_interval_sec", 1))
        self.server_timeout_sec = float(config.get("server_timeout_sec", 3600))
        self.stall_timeout_sec = float(config.get("stall_timeout_sec", 60))
        self.settle_after_finish_sec = float(config.get("settle_after_finish_sec", 2))
        self.cleanup_between_runs_sec = float(config.get("cleanup_between_runs_sec", 2))
        self.progress_every_ticks = max(1, int(config.get("progress_every_ticks", 20)))
        self.global_repeats = max(1, int(config.get("repeats", 1)))
        self.agent_command = self._normalize_command(config.get("agent_command", [sys.executable, "main.py"]))
        self.maps = [MapRun(**item) for item in config.get("maps", [])]
        if not self.maps:
            raise BenchmarkError("В конфиге нет ни одной карты для запуска")
        self.session_dir = self.output_root / f"benchmark_{self.label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.session_dir.mkdir(parents=True, exist_ok=True)
        self.runs_dir = self.session_dir / "runs"
        self.runs_dir.mkdir(parents=True, exist_ok=True)
        self.report_csv_path = self.session_dir / "benchmark_runs.csv"
        self.summary_csv_path = self.session_dir / "benchmark_summary.csv"
        self.report_md_path = self.session_dir / "benchmark_runs.md"
        self.summary_md_path = self.session_dir / "benchmark_summary.md"
        self.report_json_path = self.session_dir / "benchmark_report.json"
        self.report_xlsx_path = self.session_dir / "benchmark_report.xlsx"
        self.session_log_path = self.session_dir / "benchmark_runner.log"
        self._log_file = self.session_log_path.open("w", encoding="utf-8")

    def _detect_project_root(self) -> Path:
        env_value = os.environ.get("RRS_PROJECT_ROOT")
        if env_value:
            return Path(env_value).expanduser().resolve()
        script_path = Path(__file__).resolve()
        for candidate in [script_path.parent] + list(script_path.parents):
            if (candidate / "rcrs-server").exists() and (candidate / "newAgent").exists():
                return candidate
        return Path.cwd().resolve()

    def _resolve_path(self, value: str | Path) -> Path:
        path = Path(value).expanduser()
        if path.is_absolute():
            return path.resolve()
        return (self.project_root / path).resolve()

    def _normalize_command(self, value: Any) -> list[str]:
        if isinstance(value, str):
            return [value]
        if isinstance(value, list):
            return [str(item) for item in value]
        raise BenchmarkError("agent_command должен быть строкой или списком")

    def log(self, message: str) -> None:
        line = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"
        print(line, flush=True)
        self._log_file.write(line + "\n")
        self._log_file.flush()

    def run(self) -> None:
        run_metrics: list[RunMetrics] = []
        try:
            expanded_runs: list[tuple[int, int, MapRun]] = []
            for map_run in self.maps:
                repeats = max(1, int(map_run.repeats or self.global_repeats))
                if repeats == 1:
                    repeats = self.global_repeats
                for repeat_index in range(1, repeats + 1):
                    expanded_runs.append((len(expanded_runs) + 1, repeat_index, map_run))
            for ordinal, repeat_index, map_run in expanded_runs:
                self.log(f"[{ordinal}/{len(expanded_runs)}] {derive_map_label(map_run)} repeat={repeat_index}")
                run_metrics.append(self._run_single(map_run, repeat_index))
                time.sleep(self.cleanup_between_runs_sec)
            summaries = build_summary_rows(run_metrics)
            self._write_reports(run_metrics, summaries)
            self.log(f"Готово: {self.report_xlsx_path}")
        finally:
            self._log_file.close()

    def _run_single(self, map_run: MapRun, repeat_index: int) -> RunMetrics:
        map_dir = self._resolve_path(map_run.map_path)
        config_dir = self._resolve_path(map_run.config_path)
        map_name = derive_map_label(map_run)
        run_id = f"{map_name}__r{repeat_index:02d}"
        run_dir = self.runs_dir / run_id
        if run_dir.exists():
            shutil.rmtree(run_dir)
        run_dir.mkdir(parents=True, exist_ok=True)
        server_log_dir = run_dir / "server_logs"
        server_log_dir.mkdir(parents=True, exist_ok=True)
        server_stdout_path = run_dir / f"{map_name}__server_stdout.log"
        server_stderr_path = run_dir / f"{map_name}__server_stderr.log"
        agent_stdout_path = run_dir / f"{map_name}__agent_stdout.log"
        agent_stderr_path = run_dir / f"{map_name}__agent_stderr.log"
        copied_agent_logs_dir = run_dir / "agent_logs"
        copied_agent_logs_dir.mkdir(parents=True, exist_ok=True)
        diagnostics_path = run_dir / "diagnostics.txt"

        total_civilians, refuge_ids, building_ids = parse_map_metadata(map_dir)
        expected_timesteps = read_expected_timesteps(config_dir)
        previous_agent_snapshot = snapshot_files(self.agent_result_dir)

        env = os.environ.copy()
        env["RRS_BENCHMARK_MAP_NAME"] = map_name
        env["RRS_BENCHMARK_RUN_ID"] = run_id
        env["RRS_BENCHMARK_VARIANT"] = self.variant_label

        wall_start = time.monotonic()
        notes: list[str] = []
        max_tick_seen = 0
        next_progress_tick = self.progress_every_ticks
        last_progress_wall = time.monotonic()
        server_proc: Optional[subprocess.Popen[str]] = None
        agent_proc: Optional[subprocess.Popen[str]] = None
        stderr_reader = IncrementalLineReader(agent_stderr_path)

        try:
            server_cmd = [
                "bash",
                "./start-comprun.sh",
                "-m",
                str(map_dir),
                "-c",
                str(config_dir),
                "-l",
                str(server_log_dir),
            ]
            self.log(f"{run_id}: старт сервера")
            server_proc = self._spawn(server_cmd, self.server_scripts_dir, server_stdout_path, server_stderr_path, env)
            time.sleep(self.startup_delay_sec)
            self.log(f"{run_id}: старт агента")
            agent_proc = self._spawn(self.agent_command, self.agent_dir, agent_stdout_path, agent_stderr_path, env)
            while True:
                for line in stderr_reader.read_new_lines():
                    tick = extract_tick_from_text(line)
                    if tick is not None and tick > max_tick_seen:
                        max_tick_seen = tick
                        last_progress_wall = time.monotonic()
                        while max_tick_seen >= next_progress_tick:
                            self.log(f"{run_id}: tick={next_progress_tick}")
                            next_progress_tick += self.progress_every_ticks
                elapsed = time.monotonic() - wall_start
                server_exited = server_proc.poll() is not None if server_proc else False
                agent_exited = agent_proc.poll() is not None if agent_proc else False
                if expected_timesteps is not None and max_tick_seen >= expected_timesteps:
                    notes.append(f"timesteps_reached={expected_timesteps}")
                    time.sleep(self.settle_after_finish_sec)
                    break
                if server_exited and agent_exited:
                    notes.append("server_and_agent_exited")
                    break
                if elapsed > self.server_timeout_sec:
                    notes.append("timeout")
                    break
                if time.monotonic() - last_progress_wall > self.stall_timeout_sec:
                    notes.append(f"stalled_at_tick={max_tick_seen}")
                    break
                time.sleep(self.poll_interval_sec)
        except Exception as exc:
            notes.append(f"exception={exc.__class__.__name__}:{exc}")
            diagnostics_path.write_text(traceback.format_exc(), encoding="utf-8")
        finally:
            terminate_process(agent_proc)
            terminate_process(server_proc)
            run_kill_script(self.server_scripts_dir)
            copied = copy_new_files(self.agent_result_dir, previous_agent_snapshot, copied_agent_logs_dir, prefix=map_name)
            notes.append(f"copied_agent_logs={len(copied)}")

        wall_runtime_sec = round(time.monotonic() - wall_start, 3)
        log_bundle = collect_log_files(run_dir)
        events = parse_module_events(log_bundle)
        rescue_ticks, saved_count, rescue_source = detect_rescues_from_events(events, refuge_ids)
        if not rescue_ticks:
            text_ticks = detect_rescue_ticks_from_text(log_bundle)
            rescue_ticks = text_ticks
            saved_count = len(text_ticks)
            rescue_source = "text_logs" if text_ticks else "none"
        notes.append(f"rescue_source={rescue_source}")
        visited_buildings = detect_visited_buildings(events, building_ids)
        decision_ms = compute_tick_decision_times_ms(events)
        path_ms = compute_path_calc_times_ms(events)
        first_move_delay = compute_first_move_delay(events)
        latest_tick = max_tick_seen if max_tick_seen > 0 else detect_max_tick_from_logs(log_bundle)
        rescued_percent = round((saved_count / total_civilians) * 100.0, 2) if total_civilians > 0 else None
        explored_percent = round((len(visited_buildings) / len(building_ids)) * 100.0, 2) if building_ids else None

        return RunMetrics(
            run_id=run_id,
            variant_label=self.variant_label,
            map_name=map_name,
            map_path=str(map_dir),
            repeat_index=repeat_index,
            total_civilians=total_civilians,
            total_search_buildings=len(building_ids),
            saved_civilians=saved_count,
            first_rescue_tick=rescue_ticks[0] if rescue_ticks else None,
            last_rescue_tick=rescue_ticks[-1] if rescue_ticks else None,
            avg_ticks_between_rescues=average_intervals(rescue_ticks),
            rescued_percent=rescued_percent,
            explored_buildings_percent=explored_percent,
            avg_decision_time_ms=round_or_none(decision_ms),
            avg_path_calc_time_ms=round_or_none(path_ms),
            first_move_delay_ticks=round_or_none(first_move_delay),
            overall_ticks=latest_tick,
            wall_runtime_sec=wall_runtime_sec,
            repeatability_cv_percent=None,
            notes="; ".join(notes),
        )

    def _spawn(
        self,
        cmd: list[str],
        cwd: Path,
        stdout_path: Path,
        stderr_path: Path,
        env: dict[str, str],
    ) -> subprocess.Popen[str]:
        stdout_file = stdout_path.open("w", encoding="utf-8")
        stderr_file = stderr_path.open("w", encoding="utf-8")
        return subprocess.Popen(
            cmd,
            cwd=str(cwd),
            stdout=stdout_file,
            stderr=stderr_file,
            text=True,
            env=env,
            start_new_session=True,
        )

    def _write_reports(self, run_metrics: list[RunMetrics], summary_rows: list[SummaryRow]) -> None:
        write_csv(self.report_csv_path, [asdict(item) for item in run_metrics])
        write_csv(self.summary_csv_path, [asdict(item) for item in summary_rows])
        write_markdown(self.report_md_path, [asdict(item) for item in run_metrics])
        write_markdown(self.summary_md_path, [asdict(item) for item in summary_rows])
        write_workbook(self.report_xlsx_path, run_metrics, summary_rows)
        payload = {
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "project_root": str(self.project_root),
            "session_dir": str(self.session_dir),
            "runs": [asdict(item) for item in run_metrics],
            "summary": [asdict(item) for item in summary_rows],
            "config": self.config,
        }
        self.report_json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def derive_map_label(map_run: MapRun) -> str:
    map_path = Path(map_run.map_path)
    city = map_path.parent.name if map_path.parent else "map"
    map_leaf = map_path.name
    derived = f"{city}_{map_leaf}"
    if map_run.name and map_run.name.lower() not in {"test", "run", "map"}:
        return sanitize_label(map_run.name)
    return sanitize_label(derived)


def sanitize_label(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    with path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: empty_if_none(value) for key, value in row.items()})


def write_markdown(path: Path, rows: list[dict[str, Any]]) -> None:
    if not rows:
        return
    headers = list(rows[0].keys())
    widths = {header: len(header) for header in headers}
    prepared: list[list[str]] = []
    for row in rows:
        current = [str(empty_if_none(row.get(header, ""))) for header in headers]
        prepared.append(current)
        for header, value in zip(headers, current):
            widths[header] = max(widths[header], len(value))
    with path.open("w", encoding="utf-8") as fh:
        header_line = "| " + " | ".join(header.ljust(widths[header]) for header in headers) + " |\n"
        sep_line = "| " + " | ".join("-" * widths[header] for header in headers) + " |\n"
        fh.write(header_line)
        fh.write(sep_line)
        for row in prepared:
            fh.write("| " + " | ".join(value.ljust(widths[header]) for value, header in zip(row, headers)) + " |\n")


def write_workbook(path: Path, run_metrics: list[RunMetrics], summary_rows: list[SummaryRow]) -> None:
    wb = Workbook()
    ws_runs = wb.active
    ws_runs.title = "Runs"
    write_sheet(ws_runs, [asdict(item) for item in run_metrics])
    ws_summary = wb.create_sheet("Summary")
    write_sheet(ws_summary, [asdict(item) for item in summary_rows])
    ws_about = wb.create_sheet("About")
    ws_about["A1"] = "variant_label"
    ws_about["B1"] = run_metrics[0].variant_label if run_metrics else ""
    ws_about["A3"] = "notes"
    ws_about["B3"] = "avg_decision_time_ms и avg_path_calc_time_ms считаются по меткам wall_elapsed в логах модулей; repeatability_cv_percent заполняется только когда для одной и той же карты есть 2 и более повторов."
    style_sheet(ws_about)
    wb.save(path)


def write_sheet(ws, rows: list[dict[str, Any]]) -> None:
    if not rows:
        ws["A1"] = "Нет данных"
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([empty_if_none(row.get(header)) for header in headers])
    style_sheet(ws)


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E1F2")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 40)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = PatternFill("solid", fgColor="F7FBFF")


def collect_log_files(run_dir: Path) -> list[Path]:
    return sorted([path for path in run_dir.rglob("*") if path.is_file() and path.suffix.lower() in TEXT_FILE_SUFFIXES])


def parse_module_events(paths: list[Path]) -> list[ModuleEvent]:
    events: list[ModuleEvent] = []
    seen: set[tuple[Any, ...]] = set()
    for path in paths:
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as fh:
                for raw_line in fh:
                    line = raw_line.strip()
                    match = PREFIX_RE.match(line)
                    if not match:
                        continue
                    message = match.group("msg")
                    event_type = classify_event_type(message)
                    target_id = extract_id_after(message, "id=")
                    state = extract_state(message)
                    key = (
                        match.group("module"),
                        match.group("agent"),
                        int(match.group("tick")),
                        int(match.group("sim")),
                        float(match.group("wall")),
                        to_int(match.group("area")),
                        to_int(match.group("x")),
                        to_int(match.group("y")),
                        message,
                    )
                    if key in seen:
                        continue
                    seen.add(key)
                    events.append(
                        ModuleEvent(
                            module=match.group("module"),
                            agent=match.group("agent"),
                            tick=int(match.group("tick")),
                            sim_elapsed=int(match.group("sim")),
                            wall_elapsed_sec=float(match.group("wall")),
                            area=to_int(match.group("area")),
                            x=to_int(match.group("x")),
                            y=to_int(match.group("y")),
                            message=message,
                            source_file=path.name,
                            event_type=event_type,
                            target_id=target_id,
                            state=state,
                        )
                    )
        except Exception:
            continue
    events.sort(key=lambda item: (item.tick, item.wall_elapsed_sec, item.module, item.source_file))
    return events


def classify_event_type(message: str) -> str:
    if message.startswith("human_target_selected"):
        return "human_target_selected"
    if message.startswith("search_target_selected"):
        return "search_target_selected"
    if message.startswith("path_computed"):
        return "path_computed"
    if message.startswith("agent_position_changed"):
        return "agent_position_changed"
    if "error where=" in message:
        return "error"
    if message.startswith("module_loaded"):
        return "module_loaded"
    return "other"


def extract_id_after(text: str, token: str) -> Optional[int]:
    idx = text.find(token)
    if idx < 0:
        return None
    match = NUMBER_RE.search(text[idx + len(token):])
    if not match:
        return None
    return int(match.group(0))


def extract_state(text: str) -> Optional[str]:
    match = re.search(r"\bstate=([^\s]+)", text)
    if not match:
        return None
    return match.group(1)


def detect_rescues_from_events(events: list[ModuleEvent], refuge_ids: set[int]) -> tuple[list[int], int, str]:
    release_records: list[tuple[int, int, Optional[int], str]] = []
    carrying: dict[str, Optional[int]] = defaultdict(lambda: None)
    carrying_area: dict[str, Optional[int]] = defaultdict(lambda: None)
    carrying_tick: dict[str, Optional[int]] = defaultdict(lambda: None)

    for event in events:
        if event.event_type == "agent_position_changed":
            if carrying[event.agent] is not None and event.area is not None:
                carrying_area[event.agent] = event.area
            continue
        if event.event_type != "human_target_selected":
            continue

        current_id = carrying[event.agent]
        same_carry = event.state == "carrying" and event.target_id == current_id and current_id is not None
        if current_id is not None and not same_carry:
            release_area = carrying_area[event.agent] if carrying_area[event.agent] is not None else event.area
            release_tick = event.tick
            if release_area is not None:
                release_records.append((release_tick, current_id, release_area, event.agent))
            carrying[event.agent] = None
            carrying_area[event.agent] = None
            carrying_tick[event.agent] = None

        if event.state == "carrying" and event.target_id is not None:
            carrying[event.agent] = event.target_id
            carrying_area[event.agent] = event.area
            carrying_tick[event.agent] = event.tick

    if not release_records:
        return [], 0, "none"

    known_refuges = {area for _, _, area, _ in release_records if area in refuge_ids}
    inferred_refuges: set[int] = set()
    if not known_refuges:
        area_to_ids: dict[int, set[int]] = defaultdict(set)
        area_counts: Counter[int] = Counter()
        for _, civilian_id, area, _ in release_records:
            area_to_ids[area].add(civilian_id)
            area_counts[area] += 1
        inferred_refuges = {area for area, ids in area_to_ids.items() if len(ids) >= 2 or area_counts[area] >= 2}
        if not inferred_refuges and len(area_counts) == 1:
            inferred_refuges = set(area_counts.keys())

    active_refuges = known_refuges or inferred_refuges
    rescued_ids: set[int] = set()
    rescue_ticks: list[int] = []
    for tick, civilian_id, area, _agent in sorted(release_records):
        if area not in active_refuges:
            continue
        if civilian_id in rescued_ids:
            continue
        rescued_ids.add(civilian_id)
        rescue_ticks.append(tick)

    if rescue_ticks:
        if known_refuges:
            source = "agent_lifecycle_refuge"
        elif inferred_refuges:
            source = "agent_lifecycle_inferred_refuge"
        else:
            source = "agent_lifecycle"
        return rescue_ticks, len(rescued_ids), source
    return [], 0, "none"


def detect_rescue_ticks_from_text(paths: list[Path]) -> list[int]:
    result: set[int] = set()
    for path in paths:
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as fh:
                for line in fh:
                    lower = line.lower()
                    if not any(word in lower for word in EVENT_WORDS):
                        continue
                    tick = extract_tick_from_text(line)
                    if tick is not None:
                        result.add(tick)
        except Exception:
            continue
    return sorted(result)


def detect_visited_buildings(events: list[ModuleEvent], building_ids: set[int]) -> set[int]:
    visited: set[int] = set()
    for event in events:
        if event.event_type == "agent_position_changed" and event.area in building_ids:
            visited.add(event.area)
    return visited


def compute_tick_decision_times_ms(events: list[ModuleEvent]) -> Optional[float]:
    tick_times: dict[int, list[float]] = defaultdict(list)
    for event in events:
        if event.event_type in {"module_loaded", "error"}:
            continue
        tick_times[event.tick].append(event.wall_elapsed_sec)
    durations: list[float] = []
    for times in tick_times.values():
        if not times:
            continue
        durations.append((max(times) - min(times)) * 1000.0)
    if not durations:
        return None
    return round(sum(durations) / len(durations), 3)


def compute_path_calc_times_ms(events: list[ModuleEvent]) -> Optional[float]:
    by_tick: dict[int, list[ModuleEvent]] = defaultdict(list)
    for event in events:
        by_tick[event.tick].append(event)
    values: list[float] = []
    for tick, tick_events in by_tick.items():
        tick_events.sort(key=lambda item: item.wall_elapsed_sec)
        path_events = [event for event in tick_events if event.event_type == "path_computed"]
        if not path_events:
            continue
        path_start = path_events[0].wall_elapsed_sec
        path_end = path_events[-1].wall_elapsed_sec
        previous_events = [event for event in tick_events if event.wall_elapsed_sec <= path_start and event.event_type != "path_computed"]
        if previous_events:
            path_start = previous_events[-1].wall_elapsed_sec
        values.append(max(0.0, (path_end - path_start) * 1000.0))
    if not values:
        return None
    return round(sum(values) / len(values), 3)


def compute_first_move_delay(events: list[ModuleEvent]) -> Optional[float]:
    sorted_events = sorted(events, key=lambda item: (item.tick, item.wall_elapsed_sec))
    first_goal_tick: Optional[int] = None
    area_at_goal: Optional[int] = None
    for event in sorted_events:
        if event.event_type in {"human_target_selected", "search_target_selected", "path_computed"} and "target=None" not in event.message:
            first_goal_tick = event.tick
            area_at_goal = event.area
            break
    if first_goal_tick is None:
        return None
    for event in sorted_events:
        if event.tick < first_goal_tick:
            continue
        if event.event_type == "agent_position_changed" and event.area is not None and event.area != area_at_goal:
            return float(event.tick - first_goal_tick)
    return None


def detect_max_tick_from_logs(paths: list[Path]) -> Optional[int]:
    best: Optional[int] = None
    for path in paths:
        try:
            with path.open("r", encoding="utf-8", errors="ignore") as fh:
                for line in fh:
                    tick = extract_tick_from_text(line)
                    if tick is not None and (best is None or tick > best):
                        best = tick
        except Exception:
            continue
    return best


def average_intervals(values: list[int]) -> Optional[float]:
    if len(values) < 2:
        return None
    diffs = [values[index] - values[index - 1] for index in range(1, len(values))]
    return round(sum(diffs) / len(diffs), 3) if diffs else None


def build_summary_rows(run_metrics: list[RunMetrics]) -> list[SummaryRow]:
    groups: dict[str, list[RunMetrics]] = defaultdict(list)
    for item in run_metrics:
        groups[item.map_name].append(item)
    rows: list[SummaryRow] = []
    for map_name in sorted(groups):
        rows.append(build_summary_row(map_name, groups[map_name]))
    rows.append(build_summary_row("OVERALL", run_metrics))
    apply_repeatability_to_runs(run_metrics, groups)
    return rows


def apply_repeatability_to_runs(run_metrics: list[RunMetrics], groups: dict[str, list[RunMetrics]]) -> None:
    by_map = {map_name: coefficient_of_variation([item.rescued_percent for item in items]) for map_name, items in groups.items()}
    for item in run_metrics:
        item.repeatability_cv_percent = by_map.get(item.map_name)


def build_summary_row(group_name: str, items: list[RunMetrics]) -> SummaryRow:
    return SummaryRow(
        group_name=group_name,
        runs=len(items),
        total_civilians_avg=average_values([item.total_civilians for item in items]),
        total_search_buildings_avg=average_values([item.total_search_buildings for item in items]),
        saved_civilians_avg=average_values([item.saved_civilians for item in items]),
        first_rescue_tick_avg=average_values([item.first_rescue_tick for item in items]),
        last_rescue_tick_avg=average_values([item.last_rescue_tick for item in items]),
        avg_ticks_between_rescues_avg=average_values([item.avg_ticks_between_rescues for item in items]),
        rescued_percent_avg=average_values([item.rescued_percent for item in items]),
        explored_buildings_percent_avg=average_values([item.explored_buildings_percent for item in items]),
        avg_decision_time_ms_avg=average_values([item.avg_decision_time_ms for item in items]),
        avg_path_calc_time_ms_avg=average_values([item.avg_path_calc_time_ms for item in items]),
        first_move_delay_ticks_avg=average_values([item.first_move_delay_ticks for item in items]),
        overall_ticks_avg=average_values([item.overall_ticks for item in items]),
        wall_runtime_sec_avg=average_values([item.wall_runtime_sec for item in items]),
        repeatability_cv_percent=coefficient_of_variation([item.rescued_percent for item in items]),
        saved_civilians_std=std_values([item.saved_civilians for item in items]),
        notes="summary_by_map" if group_name != "OVERALL" else "overall_average",
    )


def average_values(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if not numeric:
        return None
    return round(sum(numeric) / len(numeric), 3)


def std_values(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if len(numeric) < 2:
        return None
    return round(statistics.pstdev(numeric), 3)


def coefficient_of_variation(values: Iterable[Optional[float]]) -> Optional[float]:
    numeric = [float(value) for value in values if value is not None]
    if len(numeric) < 2:
        return None
    mean = sum(numeric) / len(numeric)
    if abs(mean) < 1e-9:
        return None
    return round((statistics.pstdev(numeric) / mean) * 100.0, 3)


def round_or_none(value: Optional[float]) -> Optional[float]:
    return None if value is None else round(value, 3)


def empty_if_none(value: Any) -> Any:
    return "" if value is None else value


def extract_tick_from_text(line: str) -> Optional[int]:
    match = TICK_RE.search(line)
    if match:
        return int(match.group(1))
    numbers = [int(item) for item in re.findall(r"\b\d+\b", line)]
    if not numbers:
        return None
    for number in numbers:
        if 0 <= number <= 10000:
            return number
    return None


def snapshot_files(directory: Path) -> set[str]:
    if not directory.exists():
        return set()
    return {str(path.resolve()) for path in directory.rglob("*") if path.is_file()}


def copy_new_files(source_dir: Path, old_snapshot: set[str], dest_dir: Path, prefix: str) -> list[Path]:
    copied: list[Path] = []
    if not source_dir.exists():
        return copied
    for path in source_dir.rglob("*"):
        if not path.is_file():
            continue
        resolved = str(path.resolve())
        if resolved in old_snapshot:
            continue
        target = dest_dir / f"{prefix}__{path.name}"
        counter = 1
        while target.exists():
            target = dest_dir / f"{prefix}__{path.stem}_{counter}{path.suffix}"
            counter += 1
        shutil.copy2(path, target)
        copied.append(target)
    return copied


def terminate_process(proc: Optional[subprocess.Popen[str]], grace_sec: float = 8.0) -> None:
    if proc is None or proc.poll() is not None:
        return
    try:
        os.killpg(proc.pid, signal.SIGTERM)
    except Exception:
        try:
            proc.terminate()
        except Exception:
            return
    start = time.monotonic()
    while time.monotonic() - start < grace_sec:
        if proc.poll() is not None:
            return
        time.sleep(0.25)
    try:
        os.killpg(proc.pid, signal.SIGKILL)
    except Exception:
        try:
            proc.kill()
        except Exception:
            return


def run_kill_script(server_scripts_dir: Path) -> None:
    kill_path = server_scripts_dir / "kill.sh"
    if not kill_path.exists():
        return
    try:
        subprocess.run(["bash", "./kill.sh"], cwd=str(server_scripts_dir), stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=45)
    except Exception:
        pass


def parse_map_metadata(map_dir: Path) -> tuple[int, set[int], set[int]]:
    scenario_path = map_dir / "scenario.xml"
    if not scenario_path.exists():
        raise BenchmarkError(f"Не найден scenario.xml: {scenario_path}")
    root = ET.parse(scenario_path).getroot()
    refuge_ids: set[int] = set()
    civilian_ids: set[int] = set()
    synthetic_civilian_id = 1

    for elem in root.iter():
        tag = strip_ns(elem.tag).lower()
        blob = (tag + " " + " ".join(f"{k}={v}" for k, v in elem.attrib.items()) + " " + (elem.text or "")).lower()
        entity_type = (elem.attrib.get("type") or elem.attrib.get("entity_type") or "").lower()
        is_refuge = "refuge" in tag or entity_type == "refuge" or "refuge" in blob
        is_civilian = "civilian" in tag or entity_type == "civilian" or "civilian" in blob

        if is_refuge:
            entity_id = first_int_attr(elem.attrib, ["id", "entity_id", "entityId", "value", "location"])
            location = first_int_attr(elem.attrib, ["location", "position", "area", "id"])
            if location is not None:
                refuge_ids.add(location)
            elif entity_id is not None:
                refuge_ids.add(entity_id)

        if is_civilian:
            entity_id = first_int_attr(elem.attrib, ["id", "entity_id", "entityId", "value"])
            if entity_id is None:
                entity_id = synthetic_civilian_id
                synthetic_civilian_id += 1
            civilian_ids.add(entity_id)

    building_ids, refuge_from_gml = parse_buildings_and_refuges_from_map_gml(map_dir)
    refuge_ids |= refuge_from_gml
    building_ids -= refuge_ids
    return len(civilian_ids), refuge_ids, building_ids


def parse_buildings_and_refuges_from_map_gml(map_dir: Path) -> tuple[set[int], set[int]]:
    candidates = [map_dir / "map.gml"] + list(map_dir.glob("*.gml"))
    target = next((path for path in candidates if path.exists()), None)
    if target is None:
        return set(), set()
    building_ids: set[int] = set()
    refuge_ids: set[int] = set()
    try:
        root = ET.parse(target).getroot()
        for elem in root.iter():
            blob = (strip_ns(elem.tag) + " " + " ".join(f"{k}={v}" for k, v in elem.attrib.items()) + " " + (elem.text or "")).lower()
            entity_id = first_int_attr(elem.attrib, ["id", "entity_id", "entityId", "value", "gml:id", "fid"])
            if entity_id is None:
                entity_id = extract_first_int(blob)
            if entity_id is None:
                continue
            if "refuge" in blob:
                refuge_ids.add(entity_id)
                continue
            if "building" in blob or "ambulancecentre" in blob or "firestation" in blob or "policeoffice" in blob:
                building_ids.add(entity_id)
    except Exception:
        pass
    if refuge_ids or building_ids:
        return building_ids, refuge_ids
    try:
        text = target.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return set(), set()
    lines = text.splitlines()
    for idx, line in enumerate(lines):
        lower = line.lower()
        if not any(token in lower for token in ("refuge", "building", "ambulancecentre", "firestation", "policeoffice")):
            continue
        window = " ".join(lines[max(0, idx - 2): idx + 3]).lower()
        entity_id = extract_first_int(window)
        if entity_id is None:
            continue
        if "refuge" in window:
            refuge_ids.add(entity_id)
        elif any(token in window for token in ("building", "ambulancecentre", "firestation", "policeoffice")):
            building_ids.add(entity_id)
    return building_ids, refuge_ids


def read_expected_timesteps(config_dir: Path) -> Optional[int]:
    candidates = [config_dir / "kernel.cfg", config_dir / "config.cfg"] + list(config_dir.rglob("kernel.cfg"))
    for path in candidates:
        if not path.exists():
            continue
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        for line in text.splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "kernel.timesteps" in line or "kernel.timesteps=" in line:
                match = NUMBER_RE.search(line)
                if match:
                    return int(match.group(0))
    return None


def strip_ns(tag: str) -> str:
    return tag.split("}")[-1] if "}" in tag else tag


def first_int_attr(attrs: dict[str, str], keys: Iterable[str]) -> Optional[int]:
    lowered = {str(key).lower(): value for key, value in attrs.items()}
    for key in keys:
        value = lowered.get(key.lower())
        number = to_int(value)
        if number is not None:
            return number
    return None


def extract_first_int(text: str) -> Optional[int]:
    match = NUMBER_RE.search(text)
    return int(match.group(0)) if match else None


def to_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        match = NUMBER_RE.search(value)
        if match:
            return int(match.group(0))
    return None


def load_config(config_path: Path) -> dict[str, Any]:
    return json.loads(config_path.read_text(encoding="utf-8"))


def build_default_config() -> dict[str, Any]:
    return {
        "run_label": "benchmark",
        "variant_label": "baseline_variant",
        "server_scripts_dir": "rcrs-server/scripts",
        "agent_dir": "newAgent",
        "output_dir": "test",
        "agent_command": ["python", "main.py"],
        "startup_delay_sec": 5,
        "poll_interval_sec": 1,
        "server_timeout_sec": 3600,
        "stall_timeout_sec": 60,
        "settle_after_finish_sec": 2,
        "cleanup_between_runs_sec": 2,
        "progress_every_ticks": 20,
        "repeats": 1,
        "maps": [
            {"name": "berlin_map1", "map_path": "test_maps/type_map/berlin/map1", "config_path": "test_maps/config"},
            {"name": "berlin_map3", "map_path": "test_maps/type_map/berlin/map3", "config_path": "test_maps/config"},
            {"name": "paris_map2", "map_path": "test_maps/type_map/paris/map2", "config_path": "test_maps/config"},
            {"name": "paris_map4", "map_path": "test_maps/type_map/paris/map4", "config_path": "test_maps/config"},
        ],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--config", type=str, default="test/benchmark_config.json")
    parser.add_argument("--write-default-config", action="store_true")
    args = parser.parse_args()
    config_path = Path(args.config).expanduser()
    if not config_path.is_absolute():
        config_path = (Path.cwd() / config_path).resolve()
    if args.write_default_config:
        config_path.parent.mkdir(parents=True, exist_ok=True)
        config_path.write_text(json.dumps(build_default_config(), ensure_ascii=False, indent=2), encoding="utf-8")
        print(config_path)
        return 0
    if not config_path.exists():
        raise SystemExit(f"Не найден конфиг: {config_path}")
    config = load_config(config_path)
    runner = BenchmarkRunner(config)
    runner.run()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
